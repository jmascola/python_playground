import requests
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
from openpyxl.styles import Font, PatternFill, GradientFill, Alignment, Border, Side  
# importing styles for formatting the excel output @jenny

# Remotive Jobs API endpoint
jobs_url = "https://remotive.com/api/remote-jobs"

# Filters
KEYWORDS = ["python", "ai", "data"]          # keyword filter (case-insensitive)
CATEGORY_FILTER = ""                         # Double check that this doesnt throw an error, example: "Software Development" (leave "" for no filter)
MAX_DAYS_OLD = 30                            # recent posting filter (keep jobs posted within last N days)

#User can enter in their own keywords
user_keywords_input = input(
    "Enter keywords (comma-separated), or press Enter to use defaults: "
).strip()

if user_keywords_input:
    user_keywords = [
        k.strip().lower()
        for k in user_keywords_input.split(",")
        if k.strip()
    ]
else:
    user_keywords = KEYWORDS  # fallback to defaults


# Scoring Weights (should add up to 1.0)
WEIGHT_RECENCY = 0.50
WEIGHT_KEYWORDS = 0.40
WEIGHT_SALARY = 0.10

# Recency scoring rules:
# R = 1 when 0 days old
# R = 0.5 when 7 days old
# R = 0 when older than 7 days
RECENCY_CUTOFF_DAYS = 7
RECENCY_SLOPE_DENOM = 14  # makes R=0.5 at day 7 by dividing the 7/14 to get .5

OUTPUT_PATH = "week_4/final_output.xlsx"
# ----------------------------
# HELPER FUNCTIONS
# ----------------------------
#note that using """ as a multiline string to make in code documentation

#Concerting all dates to timezone.UTC to be able to calculation recency 
def parse_iso_date(date_str):
 
#Remotive publication_date is typically ISO 8601.
#Example: "2024-07-30T10:21:11+00:00"
#This returns a timezone-aware datetime, or None if it fails.
    
    if date_str is None or date_str == "":
        return None

    try:
        dt = datetime.fromisoformat(date_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except ValueError:
        return None


def days_since(posted_dt):
    
#Returns the number of whole days since posted_dt.
#If posted_dt is None, treat as old.
    
    if posted_dt is None:
        return 9999 #treating no date as super old

    now = datetime.now(timezone.utc) # using the date now in same format as converted data from API
    delta = now - posted_dt.astimezone(timezone.utc)
    return delta.days #difference in days from posted to now


def keyword_match_count(text, keywords):
    
#Counts how many keywords appear in the text (case-insensitive).
#Each keyword counts at most once.
    
    if text is None:
        text = ""
#making case insensitive by converting keywords from API to lower case 
    text_lower = text.lower() #note on putting this outside the if loop so that it lowers all keyword and all job text to lower case once
    count = 0

    for kw in keywords:
        if kw.lower() in text_lower:
            count += 1

    return count

def recency_score(days_old):
    
#Recency rules:
#- R = 1 when days_old = 0
#- R = 0.5 when days_old = 7
#- R = 0 when days_old > 7
    
    if days_old > RECENCY_CUTOFF_DAYS:
        return 0

    r = 1 - (days_old / RECENCY_SLOPE_DENOM) 
    return max(0, r)


def keyword_score(match_count, total_keywords):
    
#Normalizes keyword match count to 0-1.
    
    if total_keywords == 0:
        return 0
    return match_count / total_keywords


def salary_score(salary_value):
    
#Salary presence score: 1 if salary exists, else 0.
    
    if salary_value is None:
        return 0
    if isinstance(salary_value, str) and salary_value.strip() == "":
        return 0
    return 1


def job_score(r, k, s):

#Weighted score:
#JobScore = WEIGHT_RECENCY*R + WEIGHT_KEYWORDS*K + WEIGHT_SALARY*S

    return (WEIGHT_RECENCY * r) + (WEIGHT_KEYWORDS * k) + (WEIGHT_SALARY * s)


def passes_filters(job, keywords, category_filter, max_days_old):

#Returns True if job passes keyword, category, and recency filters.

    title = job.get("title", "")
    description = job.get("description", "")
    category = job.get("category", "")
    pub_date_str = job.get("publication_date", "")

    text = f"{title} {description}"

# Keyword filter (if keywords provided)
    if len(keywords) > 0:
        matches = keyword_match_count(text, keywords)
        if matches == 0:
            return False

# Category filter (if set)
    if category_filter != "":
        if category != category_filter:
            return False

# Recency filter (keep jobs within last max_days_old days)
    posted_dt = parse_iso_date(pub_date_str)
    d_old = days_since(posted_dt)
    if d_old > max_days_old:
        return False

    return True



# API CALL + JSON PARSING


response = requests.get(jobs_url)
print("STATUS CODE:", response.status_code)

data = json.loads(response.text)

# Remotive API typically returns jobs under the "jobs" key
jobs_list = data.get("jobs", [])
print("Total jobs pulled from API:", len(jobs_list))



# FILTER + SCORE


filtered_jobs = []

for job in jobs_list:
    # 1) pass user_keywords into the filter function
    if passes_filters(job, user_keywords, CATEGORY_FILTER, MAX_DAYS_OLD):
        title = job.get("title", "")
        company = job.get("company_name", "")
        category = job.get("category", "")
        pub_date_str = job.get("publication_date", "")
        url = job.get("url", "")
        salary = job.get("salary", "")
        description = job.get("description", "")

        posted_dt = parse_iso_date(pub_date_str)
        d_old = days_since(posted_dt)

        text = f"{title} {description}"

        # 2) use user_keywords for counting matches
        match_count = keyword_match_count(text, user_keywords)

        r = recency_score(d_old)

        # 3) use len(user_keywords) when normalizing the keyword score
        k = keyword_score(match_count, len(user_keywords))

        s = salary_score(salary)
        score = job_score(r, k, s)

        filtered_jobs.append({
            "title": title,
            "company": company,
            "category": category,
            "publication_date": pub_date_str,
            "days_old": d_old,
            "salary": salary,
            "keyword_match_count": match_count,
            "recency_score": r,
            "keyword_score": k,
            "salary_score": s,
            "job_score": score,
            "url": url
        })

print("Total jobs after filters:", len(filtered_jobs))

# Sort by job_score descending
filtered_jobs.sort(key=lambda x: x["job_score"], reverse=True)
# for each item x in the list, use its job score value as its sorting value

# EXCEL OUTPUT (OpenPyXL)
wb = Workbook()

# Jobs sheet
ws_jobs = wb.active
ws_jobs.title = "Filtered and Ranked Jobs"


# All Jobs sheet (raw API data)

ws_all = wb.create_sheet(title="All Jobs")

# Build headers from ALL keys returned by API
all_keys = set()
for job in jobs_list:
    all_keys.update(job.keys())

all_headers = sorted(all_keys)

# Write headers with styling @jenny
for col, header in enumerate(all_headers, 1):
    cell = ws_all.cell(row=1, column=col, value=header)
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)  # @jenny
    cell.fill = GradientFill(
        type="linear",
        degree=90,
        stop=["FF0097A7", "FF004D40"]
    )  # @jenny
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # @jenny

ws_all.freeze_panes = "A2"  # @jenny
ws_all.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{len(jobs_list)+1}"  # @jenny
ws_all.row_dimensions[1].height = 20  # @jenny

# Define alternating row colors @jenny
ROW_COLORS_ALL = ["FFFFFFFF", "FFECEFF1"]  # @jenny

# Write all raw jobs with alternating row styling @jenny
for row_idx, job in enumerate(jobs_list, 2):

    # Alternate row background @jenny
    row_bg = ROW_COLORS_ALL[row_idx % 2]  # @jenny

    for col_idx, key in enumerate(all_headers, 1):
        value = job.get(key, "")

        # Convert dict/list into JSON string so Excel can store it
        if isinstance(value, (dict, list)):
            value = json.dumps(value)

        cell = ws_all.cell(row=row_idx, column=col_idx, value=value)

        # Apply fill @jenny
        cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")  # @jenny
        cell.alignment = Alignment(vertical="top", wrap_text=True)  # @jenny (optional but nice for long text)

# Make it pretty: freeze header row
ws_all.freeze_panes = "A2"

# Make it pretty: simple column widths
for col in range(1, len(all_headers) + 1):
    ws_all.column_dimensions[get_column_letter(col)].width = 20

headers = [
    "Title",
    "Company",
    "Category",
    "Publication Date",
    "Days Since Posted",
    "Salary",
    "Keyword Match Count",
    "Recency Score (R)",
    "Keyword Score (K)",
    "Salary Score (S)",
    "Job Score",
    "Link"
]

# Original Header Loop - changing to add some styling to the header row @jenny
# for col, header in enumerate(headers, 1):
#     ws_jobs.cell(row=1, column=col, value=header)

for col, header in enumerate(headers, 1):
    cell = ws_jobs.cell(row=1, column=col, value=header)
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    cell.fill = GradientFill(
        type="linear",
        degree=90,
        stop=["FF0097A7", "FF004D40"]
    )
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws_jobs.freeze_panes = "A2"
ws_jobs.auto_filter.ref = f"A1:L{len(filtered_jobs)+1}"
ws_jobs.row_dimensions[1].height = 20

# Write job rows# Define alternating row colors @jenny
ROW_COLORS = ["FFFFFF", "FFECEFF1"]

for row, job in enumerate(filtered_jobs, 2):

    rank = row - 1  # define rank @jenny

    # Adding color for the top 10 list @jenny
    if rank == 1:
        row_bg = "FFFFD700"
    elif rank == 2:
        row_bg = "FFC0C0C0"
    elif rank == 3:
        row_bg = "FFCD7F32"
    elif rank <= 10:
        row_bg = "FFE0F7FA"
    else:
        row_bg = "FFECEFF1" if row % 2 == 0 else "FFFFFFFF"

    # Apply fill to entire row @jenny
    for col in range(1, len(headers) + 1):
        cell = ws_jobs.cell(row=row, column=col)
        cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")

    ws_jobs.cell(row=row, column=1, value=str(job["title"]))
    ws_jobs.cell(row=row, column=2, value=str(job["company"]))
    ws_jobs.cell(row=row, column=3, value=str(job["category"]))
    ws_jobs.cell(row=row, column=4, value=str(job["publication_date"]))
    ws_jobs.cell(row=row, column=5, value=job["days_old"])
    ws_jobs.cell(row=row, column=6, value=str(job["salary"]))
    ws_jobs.cell(row=row, column=7, value=job["keyword_match_count"])
    ws_jobs.cell(row=row, column=8, value=job["recency_score"])
    ws_jobs.cell(row=row, column=9, value=job["keyword_score"])
    ws_jobs.cell(row=row, column=10, value=job["salary_score"])
    ws_jobs.cell(row=row, column=11, value=job["job_score"])

    # Define salary cell font @jenny
    salary_cell = ws_jobs.cell(row=row, column=6)
    if job["salary"]:
        salary_cell.font = Font(bold=True, color="FF00695C")

    # Hyperlink styled @jenny
    link_cell = ws_jobs.cell(row=row, column=12, value="View Job")
    if job["url"] != "":
        link_cell.hyperlink = job["url"]
        link_cell.style = "Hyperlink"

# Make the links to the job listing actual hyperlinks, make sure its in the loop so it does it for all listings, not just the last!
    url_value = job.get("url", "")
    link_cell = ws_jobs.cell(row=row, column=12, value=str(url_value))

    if url_value:
        link_cell.hyperlink = url_value
        link_cell.style = "Hyperlink"

# Column widths (simple)
for col in range(1, len(headers) + 1):
    col_letter = get_column_letter(col)
    ws_jobs.column_dimensions[col_letter].width = 20

ws_jobs.column_dimensions["A"].width = 45  # Title wider
ws_jobs.column_dimensions["L"].width = 15  # Link


# Save workbook
wb.save(OUTPUT_PATH)
print("Saved:", OUTPUT_PATH)