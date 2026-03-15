# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)


# import the appropriate modules (you have 3)

import json
import requests
import openpyxl
from openpyxl import Workbook

character_url = "https://rickandmortyapi.com/api/character"

# set up a workbook and worksheet titled "Rick and Morty Characters"

wb = Workbook()

ws = wb.active

ws.title = "Rick and Morty Characters"

# # assign a variable 'data' with the returned GET request

response = requests.get(character_url)

json_data = response.text

data = json.loads(json_data)

print(data)

# create the appropriate headers in openpyxl for all of the keys for a single character

headers = (list(data["results"][0].keys()))

ws.append(headers)

# loop through all of the 'results' of the data to populate the rows and columns for each character
for index, header in enumerate(headers):
    print(f"index: {index}, header: {header}")
    ws.cell(row=1, column=index+1, value=header)

for row_index, character in enumerate(data["results"]):
    for col_index, content in enumerate(character.values()):
        ws.cell(row=row_index+2, column=col_index+1, value=str(content))


# NOTE: due to the headers, the rows need to be offset by one!

wb.save("week_3/spreadsheets/jmascola_w3d3_exercise.xlsx")


# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"

ws = wb.create_sheet("Rick and Morty Locations")
ws = wb.create_sheet("Rick and Morty Episodes")

# create 2 new variables for episode_url and location_url (retrieve it from the docs!)
episode_url = "https://rickandmortyapi.com/api/location"
location_url = "https://rickandmortyapi.com/api/episode"

# populate the new worksheets appropriately with all of the data!

## Rick and Morty Locations

ws = wb["Rick and Morty Locations"]

response = requests.get(location_url)

json_data = response.text

data = json.loads(json_data)

headers = (list(data["results"][0].keys()))

for index, header in enumerate(headers):
    print(f"index: {index}, header: {header}")
    ws.cell(row=1, column=index+1, value=header)

for row_index, location in enumerate(data["results"]):
    for col_index, content in enumerate(location.values()):
        ws.cell(row=row_index+2, column=col_index+1, value=str(content))

## Rick and Morty Episodes

ws = wb["Rick and Morty Episodes"]

response = requests.get(episode_url)

json_data = response.text

data = json.loads(json_data)

headers = (list(data["results"][0].keys()))

for index, header in enumerate(headers):
    print(f"index: {index}, header: {header}")
    ws.cell(row=1, column=index+1, value=header)

for row_index, episode in enumerate(data["results"]):
    for col_index, content in enumerate(episode.values()):
        ws.cell(row=row_index+2, column=col_index+1, value=str(content))


wb.save("week_3/spreadsheets/jmascola_w3d3_exercise.xlsx")
